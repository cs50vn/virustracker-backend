import sys, os, shutil, subprocess, time, sqlite3, config
from openpyxl import *


def generateSheet():
    workbook = Workbook()
    fullPath = config.dataDir + os.sep + config.schemaInputFile

    st1 = """SELECT name from sqlite_master where type= 'table' AND name != 'sqlite_sequence' """
    st2 = """SELECT * from """
    con = sqlite3.connect(fullPath)
    cursor = con.cursor()
    cursor.execute(st1)

    workbook.remove(workbook.active)

    columnCount = 1
    for row in cursor.fetchall():
        tblName = row[0]
        print("Table: " + tblName)
        #Get all column in a table
        cursor.execute(st2 + tblName)
        names = [description[0] for description in cursor.description]
        sheet = workbook.create_sheet(tblName)
        for r in names:
            sheet.cell(row=1, column=columnCount, value=r)
            columnCount += 1
        columnCount = 1

    des = config.genDataDir + os.sep + config.schemaOutputFile
    if not os.path.exists(config.genDataDir):
        os.makedirs(config.genDataDir, exist_ok=True)
    workbook.save(filename=des)


def main(argv):
    start = time.time()
    print("===========================================================")
    print("                      \033[1;32;40mBUILD TEST DATA\033[0;37;40m")
    print("===========================================================")

    print(str(argv))
    config.buildProjectPath(argv[0], argv[1], argv[2])

    generateSheet()

    elapsedTime = time.time() - start
    print("Running time: %s s" % str(elapsedTime))

if __name__ == '__main__':
    main(sys.argv[1:])